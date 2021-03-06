import datetime
from dateutil import rrule
from random import choice
import openpyxl
from openpyxl.styles import PatternFill

from django.contrib.auth.models import User, Group
from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from django.core.urlresolvers import reverse
from django.core.exceptions import ValidationError, NON_FIELD_ERRORS
from django.db import models
from django.db.models import Q, F
from django.utils import timezone
from django.utils.translation import ugettext_lazy as _
from django.utils.encoding import python_2_unicode_compatible
from django.utils.dateparse import parse_datetime, parse_time
try:
    from django.contrib.contenttypes.fields import GenericForeignKey, GenericRelation
except ImportError:
    from django.contrib.contenttypes.generic import GenericForeignKey, GenericRelation

from con_event.models import Blackout, Registrant, SKILL_LEVEL_TNG, MatchingCriteria
from rcreg_project.extras import ascii_only_no_punct
from rcreg_project.settings import BIG_BOSS_GROUP_NAME, LOWER_BOSS_GROUP_NAME
from scheduler.app_settings import DEFAULT_REG_CAP, DEFAULT_AUD_CAP
from scheduler.models import Location, Challenge, Training, INTEREST_RATING
from swingtime.conf import settings as swingtime_settings

"""Note from Dahmer:
I started off with Swingtime and changed a lot for RollerCon's specifications.
So if you're familiar with Swingtime, this may be very different.
Also, that explains style inconsistency between old code and my add-ons.
Swingtime is a great app. I just had other stuff that I wanted as well.
"""


#===============================================================================
class OccurrenceManager(models.Manager):

    use_for_related_fields = True

    #---------------------------------------------------------------------------
    def daily_occurrences(self, dt=None):
        '''
        Returns a queryset of for instances that have any overlap with a
        particular day.

        * ``dt`` may be either a datetime.datetime, datetime.date object, or
          ``None``. If ``None``, default to the current day.

        '''
        dt = dt or datetime.datetime.now()
        start = datetime.datetime(dt.year, dt.month, dt.day)
        end = start.replace(hour=23, minute=59, second=59)
        qs = self.filter(
            models.Q(
                start_time__gte=start,
                start_time__lte=end,
            ) |
            models.Q(
                end_time__gte=start,
                end_time__lte=end,
            ) |
            models.Q(
                start_time__lt=start,
                end_time__gt=end
            )
        )

        return qs

    #---------------------------------------------------------------------------
    def gather_possibles(self, con,all_act_data):
        """Makes Otto (Auto-Scheduler) run faster by moving all db queries to 1 place.
        Takes in activity list, either list of approved but unscheduled challenges or trainings
        gets all possible Occurrences for entire group, divides up by which would suit each activity
        doesn't sort by conflict yet.
        I sacrificed readability for speed. It's a mess but it's better this way.
        """

        venues = con.venue.all()
        all_locations = Location.objects.filter(venue__in=venues)

        base_q = list(Occurrence.objects.filter(
                challenge=None, training=None, start_time__gte=con.start,
                end_time__lte=con.end).select_related('location')
                )

        possibles = all_act_data.keys()

        for act in possibles:
            this_act_data = {}

            # Start gathering locaiton per activity
            if act.location_type =='Flat Track':
                if act.is_a_training():
                    act_locations = all_locations.filter(
                            venue__in=venues,
                            location_type='Flat Track',
                            location_category="Training"
                            )

                elif act.is_a_challenge():
                    # Games have to be in C1
                    if act.gametype == "6GAME" or float(act.duration) >= 1:
                        act_locations = all_locations.filter(
                                venue__in=venues, location_type='Flat Track',
                                location_category="Competition Any Length"
                                )
                    else: # Can be n C1 or C2
                        act_locations=all_locations.filter(
                                venue__in=venues, location_type='Flat Track',
                                location_category__in=[
                                        "Competition Half Length Only",
                                        "Competition Any Length"
                                        ]
                                )

            elif act.location_type == 'EITHER Flat or Banked Track':
                if act.is_a_training():
                    act_locations = all_locations.filter(
                            location_category__in=["Training","Training or Competition"],
                            venue__in=venues, location_type__in=['Flat Track', 'Banked Track']
                            )
                elif act.is_a_challenge():
                    act_locations = all_locations.filter(
                            location_category__in=["Training or Competition",
                                    "Competition Half Length Only","Competition Any Length"
                                    ],
                            venue__in=venues, location_type__in=['Flat Track', 'Banked Track']
                            )
            else:
                act_locations = all_locations.filter(
                        venue__in=venues, location_type=act.location_type
                        )

            this_act_data["locations"] = act_locations
            # End gathering locaiton per activity #

            # Start interest, activity type, per activity
            if act.interest:
                proxy_interest = act.interest
            else:
                proxy_interest = act.get_default_interest()

            if act.is_a_training():
                # Make high demand classes in low interest timeslots and vice versa
                proxy_interest = abs(6-proxy_interest)
            elif act.is_a_challenge():
                this_act_data["proxy_interest"] = proxy_interest
            duration = float(act.duration)
            dur_delta = int(duration * 60)
            this_act_data["dur_delta"] = dur_delta
            # End interest, activity type, per activity

            act_os = []
            for o in base_q:
                if ((o.interest in [proxy_interest - 1, proxy_interest, proxy_interest + 1]) and
                        (o.location in act_locations) and
                        (o.end_time == (o.start_time + datetime.timedelta(minutes=dur_delta)))
                        ):
                    act_os.append(o)


            this_act_data["act_os"] = act_os
            interestexact = []
            interestremoved = []

            for o in act_os:
                if o.interest == proxy_interest:
                    interestexact.append(o)
                else:
                    interestremoved.append(o)
            this_act_data["interestexact"] = interestexact
            this_act_data["interestremoved"] = interestremoved

            # Update dict w/ new data
            old_act_data = all_act_data.get(act)
            old_act_data.update(this_act_data)

        return all_act_data

    #---------------------------------------------------------------------------
    def sort_possibles(self, con, all_act_data,level1pairs,prefix_base):
        """Makes Otto run faster by moving all db queries. Takes in dict with
        k of activity, v list of criteria, including possible occurrences
        now need to sort by interest match, conflicts, among self and each other,
        without too maky db hits.
        """
        from swingtime.forms import L1Check  #  Avoid circular import
        figureheads = []
        participants = []
        fpks = []
        ppks = []
        busy = {}
        busy_coaching = {}

        for attr_dict in all_act_data.values():
            these_f = attr_dict.get('figureheads')
            for f in these_f:
                if f not in figureheads:
                    figureheads.append(f)
            these_p = attr_dict.get('participants')
            for p in these_p:
                if p not in participants:
                    participants.append(p)

        for f in figureheads:
            if f.pk not in fpks:
                fpks.append(f.pk)
            if f not in busy:
                busy[f] = []

        for p in participants:
            if p.pk not in ppks:
                ppks.append(p.pk)
            if p not in busy:
                busy[p] = []

        scheduled_os = (Occurrence.objects
                .filter(start_time__gte=con.start, end_time__lte=con.end)
                .exclude(training=None,challenge=None)
                .prefetch_related('training')
                .prefetch_related('training__coach__user__registrant_set')
                .prefetch_related('challenge')
                .prefetch_related('challenge__roster1__participants')
                .prefetch_related('challenge__roster2__participants')
                )

        for o in scheduled_os:
            if o.challenge:
                for roster in [o.challenge.roster1, o.challenge.roster2]:
                    for r in roster.participants.all():
                        if r in busy:
                            r_busy = busy.get(r)
                            r_busy.append(o)
                            busy[r] = list(r_busy)
                        else:
                            busy[r] = [o]

            elif o.training:
                for c in o.training.coach.all():
                    for r in c.user.registrant_set.filter(con=con):
                        if r in busy_coaching:
                            r_busy = busy_coaching.get(r)
                            r_busy.append(o)
                            busy_coaching[r] = list(r_busy)
                        else:
                            busy_coaching[r] = [o]

        # First make a separate coaching dict for conflict checks later,
        # then conbine with the regular busy dict
        for k,v in busy_coaching.iteritems():
            if k in busy:
                r_busy = busy.get(k)
                for o in v:
                    if o not in r_busy:
                        r_busy.append(o)
                busy[k] = list(r_busy)  # Don't actually think this is necessary
            else:
                busy[k] = v

        related_blackouts = Blackout.objects.filter(registrant__in=figureheads).prefetch_related('registrant')

        for b in related_blackouts:
            r_busy = busy.get(b.registrant)
            tempo = b.make_temp_o()
            r_busy.append(tempo)
            busy[b.registrant] = list(r_busy)

        avail_score_dict={}
        for act,this_act_dict in all_act_data.iteritems():
            act_p = this_act_dict.get('participants')
            act_f = this_act_dict.get('figureheads')

            level1 = []
            level15 = []
            level2 = []
            interestexact = this_act_dict.get("interestexact")
            interestremoved = this_act_dict.get("interestremoved")

            # Check to see if any participants are coaches.
            # If so, remove coach session occurrences
            coach_participants=set(act_p).intersection(set(busy_coaching.keys()))

            # Adding any coach as a figurehead makes sure they have no conflicts.
            # Doesn't distinguish betwheen whether they're coaching at the time or not.
            # Might cause a problem by being too restrictive, giving coaches too much preference.
            for c in list(coach_participants):
                if c not in act_f:
                    act_f.append(c)

            for o in interestexact:
                figurehead_intersect = o.busy_soft(act_f,busy)
                participant_intersect = o.busy_soft(act_p,busy)

                if not figurehead_intersect:
                    if not participant_intersect:
                        level1.append(o)
                    else:
                        level2.append(o)

            for o in interestremoved:
                figurehead_intersect=o.busy_soft(act_f,busy)
                participant_intersect=o.busy_soft(act_p,busy)

                if not figurehead_intersect:
                    if not participant_intersect:
                        level15.append(o)
                    else:
                        level2.append(o)

            avail_score = ((len(level1) * 100) + (len(level15) * 10) + (len(level2) * 1))
            if avail_score in avail_score_dict:
                tmp = avail_score_dict.get(avail_score)
                tmp.append(act)
            else:
                avail_score_dict[avail_score] = [act]

            this_act_dict.update({
                    "level1": level1, "level15": level15,
                    "level2": level2, "avail_score": avail_score
                    })

        taken_os = []
        ask = avail_score_dict.keys()
        ask.sort()  #  Sorting less available to more available

        for score in ask:
            act_list = avail_score_dict.get(score)
            for act in act_list:
                oselected = False
                this_act_dict = all_act_data.get(act)
                l1 = this_act_dict.get('level1')
                l15 = this_act_dict.get('level15')
                l2 = this_act_dict.get('level2')
                figs = this_act_dict.get('figureheads')
                parts = this_act_dict.get('participants')

                if len(l1) > 0:
                    while len(l1) > 0 and not oselected:
                        o = choice(l1)
                        if o not in taken_os:
                            figurehead_intersect = o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                participant_intersect=o.busy_soft(parts,busy)
                                if not participant_intersect:
                                    prefix = (prefix_base + "-%s-occurr-%s"
                                            % (str(act.pk), str(o.pk))
                                            )
                                    level1pairs[(act, o, "Perfect Match")] = L1Check(prefix=prefix)
                                    taken_os.append(o)
                                    l1.remove(o)
                                    oselected = True

                                    for l in [figs,parts]:
                                        for r in l:
                                            r_busy = busy.get(r)
                                            r_busy.append(o)
                                    break
                                else:  # If participant intersect
                                    l1.remove(o)
                                    if o not in l2:
                                        l2.append(o)
                            else:
                                l1.remove(o)
                        else:
                            l1.remove(o)

                if len(l15) > 0 and not oselected:
                    while len(l15) > 0 and not oselected:
                        o = choice(l15)
                        if o not in taken_os:
                            figurehead_intersect = o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                participant_intersect = o.busy_soft(parts,busy)
                                if not participant_intersect:
                                    prefix = (prefix_base + "-%s-occurr-%s"
                                            % (str(act.pk), str(o.pk))
                                            )
                                    level1pairs[(act, o, "+/- Interest but no Conflicts")] = L1Check(prefix=prefix)
                                    taken_os.append(o)
                                    l15.remove(o)
                                    oselected = True
                                    for l in [figs, parts]:
                                        for r in l:
                                            r_busy = busy.get(r)
                                            r_busy.append(o)
                                    break
                                else:  # If participant intersect
                                    l15.remove(o)
                                    if o not in l2:
                                        l2.append(o)
                            else:
                                l15.remove(o)
                                if not figurehead_intersect:
                                    if o not in l2:
                                        l2.append(o)
                        else:
                            l15.remove(o)

                elif len(l2) > 0 and not oselected:
                    while len(l2) > 0 and not oselected:
                        o = choice(l2)
                        if o not in taken_os:
                            figurehead_intersect = o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                prefix = (prefix_base + "-%s-occurr-%s"
                                        % (str(act.pk), str(o.pk))
                                        )
                                level1pairs[(act, o, "+/- Interest and Player Conflicts")] = L1Check(prefix=prefix)
                                taken_os.append(o)
                                l2.remove(o)
                                oselected = True

                                for l in [figs, parts]:
                                    for r in l:
                                        r_busy = busy.get(r)
                                        r_busy.append(o)
                                break
                            else:
                                l2.remove(o)
                        else:
                            l2.remove(o)

        return level1pairs


#===============================================================================
@python_2_unicode_compatible
class Occurrence(models.Model):
    '''
    Represents the start end time for a specific occurrence of a master ``Event``
    object.
    '''
    start_time = models.DateTimeField(_('start time'))
    end_time = models.DateTimeField(_('end time'))
    location = models.ForeignKey(Location, null=True, blank=True, on_delete=models.SET_NULL)
    interest = models.IntegerField(null=True, blank=True,choices=INTEREST_RATING, default=3)

    training = models.ForeignKey(Training,null=True,blank=True,on_delete=models.SET_NULL)
    challenge = models.ForeignKey(Challenge,null=True,blank=True,on_delete=models.SET_NULL)

    objects = OccurrenceManager()

    #===========================================================================
    class Meta:
        verbose_name = _('occurrence')
        verbose_name_plural = _('occurrences')
        ordering = ('start_time', 'end_time')

    #---------------------------------------------------------------------------
    def __str__(self):
        return '{}: {}'.format(self.name, self.start_time.isoformat())

    #---------------------------------------------------------------------------
    @models.permalink
    def get_absolute_url(self):
        return ('swingtime-occurrence', [str(self.id)])

    #---------------------------------------------------------------------------
    def __lt__(self, other):
        return self.start_time < other.start_time


    #---------------------------------------------------------------------------
    def get_activity(self):
        '''
        Dahmer custom. Returns training or activity related to Event.
        Currently written to ease the transition to not having Event as a model
        '''
        activity = None

        if self.training or self.challenge:
            if self.training:
                activity = self.training
            elif self.challenge:
                activity = self.challenge

        return activity
    #---------------------------------------------------------------------------
    @property
    def activity(self):

        return self.get_activity()


    #---------------------------------------------------------------------------
    @property
    def name(self):
        activity = self.get_activity()

        if activity and activity.name:
            return activity.name
        else:
            if self.interest:
                temp_name = "Empty (Interest: " + str(self.interest) + ")"
            else:
                temp_name = "Empty"
            return temp_name
    #---------------------------------------------------------------------------
    def validate_unique(self, *args, **kwargs):
        super(Occurrence, self).validate_unique(*args, **kwargs)

        if self.start_time and self.end_time and self.location:

            qs = self.__class__._default_manager.filter(
                start_time__lt=self.end_time,
                end_time__gt=self.start_time,
                location=self.location)

            if not self._state.adding and self.pk is not None:
                qs = qs.exclude(pk=self.pk)

            if qs.exists():
                raise ValidationError({NON_FIELD_ERRORS: [
                            "You can't have more than 1 occurrence in the same"
                            " place at the same time",
                            ],})

        if self.training and self.challenge:
            raise ValidationError({NON_FIELD_ERRORS: [
                    "Occurrence cannot be BOTH a Challenge and a Training",
                    ],})

    #---------------------------------------------------------------------------
    def get_endtime(self):
        """Get end time based on the train/chal it's linked to,
        and pad an extra 15 mins for 30 min chal or 30 mins for 90 min chal
        REQUIRES event and start time.
        """
        duration = 1  # Default, may be overridden
        activity = self.get_activity()
        if activity and activity.duration:
            duration = float(activity.duration)

        dur_delta = int(duration * 60)
        end_time = self.start_time + datetime.timedelta(minutes=dur_delta)

        return end_time

    #---------------------------------------------------------------------------
    def figurehead_conflict(self):
        """Checks to see if any figureheads (coach, captain) are participating
        in other occurrances at same time. If so, returns dict with
        activity k, skater list v, but has no teeth, can be overridden,
        is just an FYI warning.
        """

        activity = self.get_activity()
        figureheads = []
        conflict_dict = {}
        if activity:
            # Figureheads is for getting blackouts
            figureheads = activity.get_figurehead_registrants()
        else:
            figureheads = []

        concurrent = list(Occurrence.objects
                .filter(start_time__lt=(self.end_time + datetime.timedelta(minutes=30)),
                        end_time__gt=(self.start_time - datetime.timedelta(minutes=30)))
                .exclude(pk=self.pk)
                .select_related('challenge')
                .select_related('training')
                )

        for o in concurrent:
            event_activity = o.get_activity()
            if event_activity:  # Could be an empty timeslot
                event_part = event_activity.participating_in()
                conflict = set(figureheads).intersection(event_part)
                if len( conflict ) > 0:
                    conflict_dict[event_activity]=list(conflict)

        if len(conflict_dict) > 0:
            return conflict_dict
        else:
            return None
    #---------------------------------------------------------------------------
    def os_soft_intersect(self,o2):

        if ((o2.start_time < (self.end_time + datetime.timedelta(minutes=30))) and
                (o2.end_time > (self.start_time - datetime.timedelta(minutes=30)))
                ):

            return True
        else:
            return False
    #---------------------------------------------------------------------------
    def os_hard_intersect(self,o2):
        if (o2.start_time < self.end_time) and (o2.end_time > self.start_time):
            return True
        else:
            return False
    #-------------------------------------------------------------------------------
    def busy_soft(self,participant_list,busy_dict):
        """takes in list of relevant registrant, dict w/ registrant as key,
        list of occurrences reg is in as v.
        checks to see if reg is busy/ w/ soft intersection.
        """

        intersection = False
        i = 0

        while not intersection and i < len(participant_list):
            for f in participant_list:
                busy_list = busy_dict.get(f)
                for o2 in busy_list:
                    if self.os_soft_intersect(o2):
                        intersection = True
                        break
                i += 1
                if intersection:
                    break
            if intersection:
                break

        return intersection

    #-------------------------------------------------------------------------------
    def participant_conflict(self):
        """Checks to see if any participants are participating in other
        occurrances at same time. If so, returns dict w/  activity k, skater list v,
        but has no teeth, can be overridden, is just an FYI warning.
        """

        activity = self.get_activity()
        figureheads = []
        conflict_dict = {}

        if activity:
            occur_part = activity.participating_in()
        else:
            occur_part=[]

        concurrent = (Occurrence.objects
                .filter(start_time__lt=(self.end_time + datetime.timedelta(minutes=30)),
                        end_time__gt=(self.start_time - datetime.timedelta(minutes=30)))
                .exclude(pk=self.pk)
                )

        for o in concurrent:
            event_activity = o.get_activity()
            if event_activity:  # Could be an empty timeslot
                event_part = event_activity.participating_in()
                inter = set(occur_part).intersection(event_part)
                if len(inter) > 0:
                    conflict_dict[event_activity] = list(inter)
        if len(conflict_dict) > 0:
            return conflict_dict
        else:
            return None

    #-------------------------------------------------------------------------------
    def blackout_conflict(self):
        """Checks to see if any figureheads (coach, captain) have blackouts during
        Occurrance time. If so, returns dict w/  blackout k, skater list v,
        but has no teeth, can be overridden, is just an FYI warning.
        """

        activity = self.get_activity()
        figureheads = []
        conflict_dict = {}
        daypart = []

        if activity:
            figureheads = activity.get_figurehead_registrants()
        else:
            figureheads = []

        odate = self.start_time.date()
        if self.start_time.time() >= parse_time('12:00:00'):
            daypart.append("PM")
        else:
            daypart.append("AM")
        if (self.end_time.time() >= parse_time('12:00:00')) and ("PM" not in daypart):
            daypart.append("PM")

        for f in figureheads:
            potential_bouts = Blackout.objects.filter(registrant=f, date=odate, ampm__in=daypart)
            if len(list(potential_bouts)) > 0:
                conflict_dict[f] = list(potential_bouts)

        if len(conflict_dict) > 0:
            return conflict_dict
        else:
            return None
    #---------------------------------------------------------------------------
    def get_add_url(self):
        """Creates string of URL to call add-event page.
        Like a DIY get absolute url"""

        dstr_str = self.start_time.isoformat()

        url_str = '/events/add/?dtstart=%s&location=%s' % (dstr_str, str(self.location.pk))

        if self.training:
            url_str += "&training=%s" % (str(self.training.pk))
        if self.challenge:
            url_str += "&challenge=%s" % (str(self.challenge.pk))

        return url_str

    #---------------------------------------------------------------------------
    def can_add_sk8ers(self):
        """Returns list of Users that can edit Roster
        this is for adding/removing roster participants,
        they are only true in activity.editable_by. This is Bosses and Volunteers.
        """
        allowed_editors = list(User.objects
                .filter(groups__name__in=['Volunteer', BIG_BOSS_GROUP_NAME, LOWER_BOSS_GROUP_NAME
                ]))

        return allowed_editors

    #---------------------------------------------------------------------------
    def excel_backup(self):
        """Writes/returns xlsx file of relevant data, to be gathered ahead
        of the con and be used as a Plan B, or for people who don't want to use
        the site, I guess (I think redundant, Ivanna insists).
        """

        wb = None

        if self.training or self.challenge:

            timestr = self.start_time.strftime('%H %M %p ')
            xlfilename = timestr + (ascii_only_no_punct(self.name)) + ".xlsx"

            wb = openpyxl.Workbook()
            sheet = wb.active

            if self.training:

                regros, rcreated = TrainingRoster.objects.get_or_create(registered=self)

                if regros.intl:
                    splitxlfilename = xlfilename.split(".")
                    newname = splitxlfilename[0] + " INTL"
                    xlfilename = ''.join([newname,splitxlfilename[1]])

                sheet["E3"].value = "Printed: %s" % (timezone.now().strftime('%m %d %Y'))
                sheet["A1"].value = self.training.name
                sheet["A2"].value = self.training.figurehead_display
                sheet["A3"].value = self.start_time.strftime('%H %M %p, %m-%d-%Y')
                sheet["A4"].value = self.location.name

                if regros.intl:
                    sheet["B6"].value = "INTL ONLY Registration Roster"
                else:
                    sheet["B6"].value = "Registration Roster"
                sheet["A7"].value = "Skill:"
                sheet["B7"].value = self.training.skill_display()
                sheet["A8"].value = "Pass:"
                sheet["B8"].value = self.training.passes_str()
                sheet["B10"].value = "Name"

                if regros.intl:
                    sheet["F6"].value = ("Auditing Roster (non-INTL skaters go "
                            "here, maybe coach will let them skate.)")
                else:
                    sheet["F6"].value = "Auditing Roster"

                sheet["E7"].value = "Skill:"
                sheet["F7"].value = "ABCD"
                sheet["E8"].value = "Pass:"
                sheet["F8"].value = "MVP, Skater, Offskate"
                sheet["F10"].value = "Name"

                starti = 11
                rno = int(1)
                rmaxcap = regros.get_maxcap()
                for r in range(1, (rmaxcap + 1)):
                    sheet["A" + str(starti)].value = str(rno) + "."
                    rno += 1
                    starti += 1

                starti = 11
                rno = int(1)
                audros, acreated = TrainingRoster.objects.get_or_create(auditing=self)
                amaxcap = audros.get_maxcap()
                for r in range(1, (amaxcap + 1)):
                    sheet["E" + str(starti)].value = str(rno) + "."
                    rno += 1
                    starti += 1

            elif self.challenge:

                sheet["A1"].value = self.challenge.data_title
                sheet["A2"].value = self.location.abbrv
                sheet["B2"].value = self.start_time.strftime('%H %M %p, %m-%d-%Y')

                sheet["D2"].value = "Printed:"
                sheet["E2"].value = timezone.now().strftime('%m %d %Y')

                if self.challenge.communication:
                    sheet["A4"].value = "ATTN:"
                    sheet["A4"].fill = PatternFill(start_color='FFFF0000', end_color="FFFFFF00", fill_type='solid')
                    sheet.merge_cells('B4:G4')
                    sheet["B4"].value = self.challenge.communication
                    sheet.row_dimensions[4].height=(12 * len( self.challenge.communication.splitlines() ) )

                sheet["A6"].value = "TEAM"
                sheet["B6"].value = self.challenge.roster1.name
                sheet["A7"].value = "COLOR"
                sheet["B7"].value = self.challenge.roster1.color
                sheet["A8"].value = "CAPTAIN"
                sheet["B8"].value = self.challenge.roster1.captain.name
                if self.challenge.roster1.captain.email:
                    sheet["C8"].value = self.challenge.roster1.captain.email
                sheet["A9"].value = "SKILL"
                sheet["B9"].value = self.challenge.roster1.skill_display()
                sheet["A10"].value = "GENDER"
                sheet["B10"].value = self.challenge.roster1.gender_text()

                sheet["E6"].value = "TEAM"
                sheet["E7"].value = "COLOR"
                sheet["F6"].value = self.challenge.roster2.name
                sheet["F7"].value = self.challenge.roster2.color
                sheet["E8"].value = "CAPTAIN"
                sheet["F8"].value = self.challenge.roster2.captain.name
                if self.challenge.roster2.captain.email:
                    sheet["G8"].value = self.challenge.roster2.captain.email

                sheet["E9"].value = "SKILL"
                sheet["F9"].value = self.challenge.roster2.skill_display()
                sheet["E10"].value = "GENDER"
                sheet["F10"].value = self.challenge.roster2.gender_text()

                sheet["A12"].value = "# of players"
                sheet["E12"].value = "# of players"
                sheet["B12"].value = "Skater #"
                sheet["F12"].value = "Skater #"
                sheet["C12"].value = "Skater Name"
                sheet["G12"].value = "Skater Name"

                starti = 13
                rno = int(1)
                r1 = list(self.challenge.roster1.participants.all())
                r1.sort(key=lambda x: x.sk8number)
                for r in r1:
                    if r == self.challenge.roster1.captain:
                        if r.sk8name:
                            name = r.sk8name+" (Captain)"
                        else:
                            name = "(Captain)"
                    else:
                        name = r.sk8name
                    sheet["A" + str(starti)].value = str(rno) + "."
                    sheet["B" + str(starti)].value = r.sk8number
                    sheet["C" + str(starti)].value = name
                    rno += 1
                    starti += 1

                starti = 13
                rno = int(1)
                r2 = list(self.challenge.roster2.participants.all())
                r2.sort(key=lambda x: x.sk8number)
                for r in r2:
                    if r == self.challenge.roster2.captain:
                        if r.sk8name:
                            name = r.sk8name+" (Captain)"
                        else:
                            name = "(Captain)"
                    else:
                        name = r.sk8name
                    sheet["E" + str(starti)].value = str(rno) + "."
                    sheet["F" + str(starti)].value = r.sk8number
                    sheet["G" + str(starti)].value = name
                    rno += 1
                    starti += 1

        return wb, xlfilename

#===============================================================================
class TrainingRoster(MatchingCriteria):
    """Used for Registration and Auditing roster for training Occurrences.
    Can be made in the Admin if made INTL, or made using get_or_create in
    register_training view. Otherwise, not necessary yet.
    """

    # From Matching_criteria: gender, con, intl, skill
    cap = models.IntegerField(null=True, blank=True)
    participants = models.ManyToManyField(Registrant, blank=True)

    # Reminder: can have 1 of these but not both.
    registered = models.OneToOneField("Occurrence", null=True, blank=True, related_name="registered")
    auditing = models.OneToOneField("Occurrence", null=True, blank=True, related_name="auditing")
    #---------------------------------------------------------------------------
    def __unicode__(self):
        return self.name
    #---------------------------------------------------------------------------
    class Meta:
        ordering=('registered__start_time', 'auditing__start_time',
                'registered__training__name', 'auditing__training__name'
                )
    #---------------------------------------------------------------------------
    @property
    def name(self):
        basename = ""
        if self.registered:
            if self.intl:
                basename += "INTL "
            basename += ("%s %s (REGISTERED)" %
                    (self.registered.name,
                    self.registered.start_time.strftime("%a %B %d %I:%-M %p")
                    ))
        elif self.auditing:
            basename += ("%s %s (AUDITING)" %
                    (self.auditing.name,
                    self.auditing.start_time.strftime("%a %B %d %I:%-M %p")
                    ))
        else:
            basename += "Error. Training Roster without a Training"

        return basename
    #---------------------------------------------------------------------------
    def validate_unique(self, *args, **kwargs):
        super(TrainingRoster, self).validate_unique(*args, **kwargs)

        if self.registered and self.auditing:
            raise ValidationError({
                NON_FIELD_ERRORS: ["Roster cannot be both Registered & Auditing",],})

        if not self.registered and not self.auditing:
            raise ValidationError({
                NON_FIELD_ERRORS: ["Please choose a Training Occurrence",],})
    #---------------------------------------------------------------------------
    def intls_allowed(self):
        if self.intl is True:
            allowed = [True]
        else:
            allowed = [True,False,None]
        return allowed
    #---------------------------------------------------------------------------
    def can_register_at(self):
        """Returns datetime registration for class starts
        both for displaying that time and checking if now is past that time"""

        can_reg = None
        regtimes = []

        if self.registered or self.auditing:
            if self.registered:
                o = self.registered
            elif self.auditing:
                o = self.auditing
            con = o.training.con

            ostart = datetime.time(hour=o.start_time.hour,minute=o.start_time.minute)
            if ostart <= con.morning_class_cutoff:
                # if class starts early enough in the morning
                yday = datetime.timedelta(days=1)
                startday = o.start_time.date() - yday
                tempdt = datetime.datetime(
                        year=startday.year,
                        month=startday.month,
                        day=startday.day,
                        hour=con.dayb4signup_start.hour,
                        minute=con.dayb4signup_start.minute
                        )
                regtimes.append(tempdt )

            # otherwise this is time - hoursb4signup
            # calculates both just in case timezone doesn't work
            # and they do something like 48 hours before class time or something
            b4signup = datetime.timedelta(hours=float(con.hoursb4signup))
            regtime = o.start_time - b4signup
            tempdt2 = datetime.datetime(
                    year=o.start_time.year,
                    month=o.start_time.month,
                    day=o.start_time.day,
                    hour=regtime.hour,
                    minute=regtime.minute
                    )
            regtimes.append(tempdt2)

            if len(regtimes) > 0:
                can_reg = min(regtimes)

        return can_reg

    #---------------------------------------------------------------------------
    def can_register(self):
        """Returns true if registration window is open, False if not.
        Will be determined by 2(?) hour window before class starts"""
        can_reg = self.can_register_at()
        now = timezone.now()

        if can_reg and now >= can_reg:
            return True
        else:
            return False
    #---------------------------------------------------------------------------
    def get_maxcap(self):
        """Gets maximum number of participants for triningroster.
        PRIORITY ORDER: self.cap for TrainingRoster, use that.
        If not, then if Training specifically has a regcap or audcap.
        If not, general default cap listed in Scheduler.models.
        If this is the auditing roster of an INTL training, allows the audit cap
        to be general training defaults-number of people registered.
        That is so coaches can have a larger audit roster in empty INTL classes.
        LOOPHOLE: people w/out an MVP pass can sign up to audit an INTL class
        and then be allowed in to participate.
        I think it'll take people a long time to figure that out, if they ever do.
        """

        intl = False

        if self.registered and self.registered.training:
            #If this is a registration training roster

            #get regcap
            if self.cap:
                regcap = self.cap
            elif self.registered.training.regcap:
                regcap = self.registered.training.regcap
            else:
                regcap = DEFAULT_REG_CAP

            if self.intl:
                intl = True

                #get audcap. Only need if is intl
                audcap = DEFAULT_AUD_CAP
                # audcap might get overwritten, just here so I don't need to do 2 elses
                if self.registered.auditing and self.registered.auditing.cap:
                    audcap = self.registered.auditing.cap
                elif self.registered.training and self.registered.training.audcap:
                    audcap = self.registered.training.audcap
            else:  # If not intl
                maxcap = regcap

        elif self.auditing and self.auditing.training:
            # If this is the auditing trainingroster
            # Need to get audcap even if is INTL and gets overridden, bc is used in equation
            if self.cap:
                audcap = self.cap
            elif self.auditing.training.audcap:
                audcap = self.auditing.training.audcap
            else:
                audcap = DEFAULT_AUD_CAP

            if self.auditing.registered.intl:
                intl = True

                regcap = DEFAULT_REG_CAP
                # regcap might get overwritten, just here so I don't need ot do 2 elses
                if self.auditing.registered and self.auditing.registered.cap:
                    regcap = self.auditing.registered.cap
                elif self.auditing.training and self.auditing.training.regcap:
                    regcap = self.auditingtraining.regcap
            else:
                maxcap = audcap

        if intl:
            if self.registered and self.registered.training:
            #If this is a registration training roster

            #what to do if is registration
                if self.registered.auditing:
                    audsk8 = self.registered.auditing.participants.count()
                else:
                    audsk8 = 0

                intlborrow = audsk8-audcap
                # How mant people were borrowed from unfulfilled regcap?
                if intlborrow > 0:
                    #How many people are allowed minus people borrowed from reg roster for audit
                    maxcap = regcap - intlborrow
                else:
                    maxcap = regcap

            elif self.auditing and self.auditing.training:
                # If this is the auditing trainingroster
                if self.auditing.registered:
                    regsk8 = self.auditing.registered.participants.count()
                else:
                    regsk8 = 0

                maxcap = ((regcap - regsk8) + audcap)

        return maxcap

    #---------------------------------------------------------------------------
    def spacea(self):
        """gets maxcap (see above), checks is participants are fewer"""
        maxcap = self.get_maxcap()
        spacea = maxcap - self.participants.count()

        if spacea > 0:
            return spacea
        else:
            return False
    #---------------------------------------------------------------------------
    def editable_by(self):
        """Returns list of Users that can edit Roster.
        For adding/removing roster participants, so coaches actually don't have
        this permission,coaches are only true in activity.editable_by.
        This is Bosses and Volunteers."""

        allowed_editors = list(User.objects.filter(
                groups__name__in=[
                        'Volunteer', BIG_BOSS_GROUP_NAME, LOWER_BOSS_GROUP_NAME
                        ]
                ))

        return allowed_editors
